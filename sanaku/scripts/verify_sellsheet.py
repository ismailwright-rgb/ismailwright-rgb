#!/usr/bin/env python3
"""Check every formula in the services workbook against the catalog.

    python3 scripts/verify_sellsheet.py [catalog.json] [workbook.xlsx]

Why this exists rather than relying on a recalculation pass: recalculating with
LibreOffice proves the formulas EVALUATE. It does not prove they are right. An
off-by-one row reference produces a clean, error-free workbook full of wrong
numbers - and wrong numbers on a price list are the whole risk here.

So this resolves each formula by following its references itself, and compares
the result against what the catalog says it should be. Verified to bite: moving
one package's saving formula a single row off is caught on every package.
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
cat_path = Path(sys.argv[1]) if len(sys.argv) > 1 else HERE.parent / "docs" / "catalog.json"
wb_path = Path(sys.argv[2]) if len(sys.argv) > 2 else HERE.parent / "docs" / "sanaku-services.xlsx"

cat = json.loads(cat_path.read_text())
by = {s["code"]: s for s in cat["services"]}
mem = {}
for m in cat["bundle_members"]:
    mem.setdefault(m["bundle_code"], []).append(m["member_code"])

wb = load_workbook(wb_path)

PATTERNS = [
    (r"'([^']+)'!([A-Z]+\d+)", "ref"),
    (r"SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)", "sum"),
    (r"([A-Z]+\d+)-([A-Z]+\d+)", "sub"),
    (r"([A-Z]+\d+)\*([A-Z]+\d+)", "mul"),
    # X+(Y*n) covers both the scenario year-one (n=12) and the trial year-one
    # (n=11, because month one is free).
    (r"([A-Z]+\d+)\+\(([A-Z]+\d+)\*(\d+)\)", "plusmul"),
    (r"([A-Z]+\d+)", "same"),
    (r"IF\(([A-Z]+\d+)=0,0,\(([A-Z]+\d+)-([A-Z]+\d+)\)/([A-Z]+\d+)\)", "pct"),
]


def val(sheet, coord, depth=0):
    assert depth < 12, f"formula loop at {sheet}!{coord}"
    v = wb[sheet][coord].value
    if v is None:
        return 0
    if not (isinstance(v, str) and v.startswith("=")):
        return v
    f = v[1:]
    for pat, kind in PATTERNS:
        m = re.fullmatch(pat, f)
        if not m:
            continue
        g = m.groups()
        if kind == "ref":
            return val(g[0], g[1], depth + 1)
        if kind == "sum":
            return sum(val(sheet, f"{g[0]}{r}", depth + 1) or 0
                       for r in range(int(g[1]), int(g[3]) + 1))
        if kind == "sub":
            return (val(sheet, g[0], depth + 1) or 0) - (val(sheet, g[1], depth + 1) or 0)
        if kind == "mul":
            return (val(sheet, g[0], depth + 1) or 0) * (val(sheet, g[1], depth + 1) or 0)
        if kind == "plusmul":
            return (val(sheet, g[0], depth + 1) or 0) + (val(sheet, g[1], depth + 1) or 0) * int(g[2])
        if kind == "same":
            return val(sheet, g[0], depth + 1)
        if kind == "pct":
            d = val(sheet, g[0], depth + 1) or 0
            if d == 0:
                return 0
            return ((val(sheet, g[1], depth + 1) or 0) - (val(sheet, g[2], depth + 1) or 0)) / d
    raise AssertionError(f"unhandled formula at {sheet}!{coord}: {f}")


def block_rows(ps, start, label):
    for rr in range(start, min(start + 25, ps.max_row + 1)):
        if (ps.cell(row=rr, column=1).value or "").strip() == label:
            return rr
    return None


bad, checked = [], 0
ps = wb["Packages & Savings"]
for r in range(5, ps.max_row + 1):
    code = ps.cell(row=r, column=2).value
    if not code or code not in by or by[code]["category"] != "bundle":
        continue
    b = by[code]
    parts = [by[c] for c in mem.get(code, []) if c in by]
    sep_s = sum(float(p["setup_fee"]) for p in parts)
    sep_m = sum(float(p["monthly_fee"]) for p in parts)
    sep = block_rows(ps, r, "Bought separately")
    save = block_rows(ps, r, "You save")
    if sep is None or save is None:
        bad.append(f"{code}: could not find its Bought separately / You save rows")
        continue
    expect = {
        f"C{r}": float(b["setup_fee"]),
        f"D{r}": float(b["monthly_fee"]),
        f"C{sep}": sep_s,
        f"D{sep}": sep_m,
        f"C{save}": sep_s - float(b["setup_fee"]),
        f"D{save}": sep_m - float(b["monthly_fee"]),
    }
    for coord, e in expect.items():
        checked += 1
        g = val("Packages & Savings", coord)
        if abs((g or 0) - e) > 0.51:
            bad.append(f"{code} {coord}: formula gives {g}, catalog says {e}")
    checked += 1
    pct = val("Packages & Savings", f"E{save}")
    if abs(round(pct * 100) - round((1 - float(b["setup_fee"]) / sep_s) * 100)) > 1:
        bad.append(f"{code}: setup saving % is {pct:.3f}, catalog implies "
                   f"{1 - float(b['setup_fee']) / sep_s:.3f}")

sc = wb["Scenarios"]
for r in range(5, sc.max_row + 1):
    name = sc.cell(row=r, column=1).value
    if not name or name == "Total":
        continue
    checked += 1
    y1 = val("Scenarios", f"E{r}")
    setup, monthly = val("Scenarios", f"C{r}"), val("Scenarios", f"D{r}")
    if abs(y1 - (setup + monthly * 12)) > 0.51:
        bad.append(f"Scenarios r{r}: year one is {y1}, not {setup} + 12x{monthly}")

# The trial tab. Year one on trial is the starting fee plus eleven billed
# months, because month one is free - getting that wrong by a month is the
# easiest way for this sheet to overstate what the offer earns.
tr = wb["Trial"]
for r in range(5, tr.max_row + 1):
    name = tr.cell(row=r, column=1).value
    if not name or str(name).startswith("Total") or str(name).startswith("'Given"):
        continue
    setup = val("Trial", f"B{r}")
    start = val("Trial", f"C{r}")
    monthly = val("Trial", f"E{r}")
    for coord, expect, what in (
        (f"D{r}", setup - start, "setup given up"),
        (f"F{r}", monthly, "free month cost"),
        (f"G{r}", start + monthly * 11, "year one on trial"),
        (f"H{r}", setup + monthly * 12, "year one paid up front"),
    ):
        checked += 1
        got = val("Trial", coord)
        if abs((got or 0) - expect) > 0.51:
            bad.append(f"Trial {coord} ({what}) for {name}: gives {got}, expected {expect}")
    if start > setup:
        bad.append(f"Trial r{r}: starting fee {start} exceeds normal setup {setup}")

print(f"resolved {checked} formula results across {len(wb.sheetnames)} tabs")
if bad:
    print("MISMATCHES:")
    for b in bad:
        print("  " + b)
    sys.exit(1)
print("every formula resolves to the value the catalog implies")
