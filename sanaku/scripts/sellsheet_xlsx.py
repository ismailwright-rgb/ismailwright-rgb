#!/usr/bin/env python3
"""Build the services workbook from the same catalog the sell sheet reads.

    python3 scripts/sellsheet_xlsx.py [catalog.json] [out.xlsx]

Every price is written as a value read from the catalog; every total, saving and
scenario is written as a FORMULA referring to those cells. Change a price on the
All Services tab and the packages, savings and scenarios all move with it -
which is the point of a spreadsheet over a printout.

Revenue only. There is no cost data anywhere in the system - no Twilio, VAPI or
OpenRouter spend is recorded against a service - so a margin column here would
be invented, and an invented margin is worse than none.
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
catalog_path = Path(sys.argv[1]) if len(sys.argv) > 1 else HERE.parent / "docs" / "catalog.json"
out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else HERE.parent / "docs" / "sanaku-services.xlsx"

catalog = json.loads(catalog_path.read_text())
services = [s for s in catalog["services"] if s.get("active") is not False]
services.sort(key=lambda s: (s["sort"], s["code"]))
by_code = {s["code"]: s for s in services}

members = {}
for m in catalog.get("bundle_members", []):
    members.setdefault(m["bundle_code"], []).append(m["member_code"])

# ---------------------------------------------------------------------------
# House style. Arial throughout; blue for a figure you may edit, black for
# anything the sheet works out for itself.
# ---------------------------------------------------------------------------
FONT = "Arial"
H1 = Font(name=FONT, size=14, bold=True)
HDR = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY = Font(name=FONT, size=10)
INPUT = Font(name=FONT, size=10, color="0000FF")
DERIVED = Font(name=FONT, size=10, bold=True)
MUTED = Font(name=FONT, size=9, color="666666")
HDR_FILL = PatternFill("solid", fgColor="0D6B42")     # the product's own green
NOTE_FILL = PatternFill("solid", fgColor="FFF9E6")
THIN = Side(style="thin", color="D9D9D9")
BOX = Border(bottom=THIN)
MONEY = '$#,##0;($#,##0);-'
MONEY2 = '$#,##0.00;($#,##0.00);-'
PCT = '0.0%'

wb = Workbook()


def header(ws, row, labels, widths=None):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font, c.fill = HDR, HDR_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


STATUS_WORD = {"live": "Available now", "in_build": "In development", "planned": "Planned"}

# ---------------------------------------------------------------------------
# 1. All Services - the single source every other tab points at.
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "All Services"
ws["A1"] = "Sanaku — every service in the catalog"
ws["A1"].font = H1
ws["A2"] = ("Blue figures are the catalog prices — edit these and every other tab follows. "
            "Generated from docs/catalog.json; re-pull it after changing a price in Supabase.")
ws["A2"].font = MUTED
ws.merge_cells("A2:J2")

COLS = ["Code", "Service", "Type", "Verticals", "Setup", "Monthly",
        "Included", "Unit", "Overage", "Per qualified lead", "BAA required", "Status"]
header(ws, 4, COLS, [26, 30, 10, 30, 11, 11, 10, 10, 10, 17, 12, 16])

row_of = {}
r = 5
for s in services:
    row_of[s["code"]] = r
    ws.cell(row=r, column=1, value=s["code"]).font = BODY
    ws.cell(row=r, column=2, value=s["name"]).font = BODY
    ws.cell(row=r, column=3, value=s["category"]).font = BODY
    ws.cell(row=r, column=4, value=", ".join(s["allowed_verticals"])).font = BODY
    for col, key, fmt in ((5, "setup_fee", MONEY), (6, "monthly_fee", MONEY)):
        c = ws.cell(row=r, column=col, value=float(s[key]))
        c.font, c.number_format = INPUT, fmt
    c = ws.cell(row=r, column=7, value=s["included_minutes"])
    c.font = INPUT
    ws.cell(row=r, column=8, value=s.get("unit_label") or "min").font = BODY
    c = ws.cell(row=r, column=9, value=float(s["overage_per_minute"]) if s["overage_per_minute"] is not None else None)
    c.font, c.number_format = INPUT, MONEY2
    c = ws.cell(row=r, column=10, value=float(s["per_lead_fee"]) if s["per_lead_fee"] is not None else None)
    c.font, c.number_format = INPUT, MONEY
    ws.cell(row=r, column=11, value="Yes" if s["requires_baa"] else "").font = BODY
    ws.cell(row=r, column=12, value=STATUS_WORD.get(s["build_status"], s["build_status"])).font = BODY
    for col in range(1, 13):
        ws.cell(row=r, column=col).border = BOX
    r += 1

ws.cell(row=r + 1, column=1,
        value="Healthcare rows carry no per-lead fee by design — per-patient compensation can "
              "implicate anti-kickback and fee-splitting rules. The database constraint "
              "sanaku_addons_pricing_legal rejects one, so it cannot be entered by mistake."
        ).font = MUTED
ws.cell(row=r + 1, column=1).fill = NOTE_FILL

# ---------------------------------------------------------------------------
# 2. Packages — every figure a formula pointing back at All Services.
# ---------------------------------------------------------------------------
ps = wb.create_sheet("Packages & Savings")
ps["A1"] = "What each package contains, and what it saves"
ps["A1"].font = H1
ps["A2"] = ("Every number on this tab is a formula reading All Services. Change a member's "
            "price there and the package saving here recalculates.")
ps["A2"].font = MUTED
ps.merge_cells("A2:G2")

header(ps, 4, ["Package / part", "Code", "Setup", "Monthly", "Note", "", ""],
       [34, 26, 13, 13, 40, 4, 4])

pr = 5
bundles = [s for s in services if s["category"] == "bundle"]
for b in bundles:
    parts = sorted((by_code[c] for c in members.get(b["code"], []) if c in by_code),
                   key=lambda x: (x["sort"], x["name"]))
    if not parts:
        continue
    bundle_row = pr          # the package's own price, for the saving below
    c = ps.cell(row=pr, column=1, value=b["name"] + "  —  " + ", ".join(b["allowed_verticals"]))
    c.font = Font(name=FONT, size=11, bold=True)
    ps.cell(row=pr, column=2, value=b["code"]).font = MUTED
    bref = row_of[b["code"]]
    for col, src in ((3, "E"), (4, "F")):
        cc = ps.cell(row=pr, column=col, value=f"='All Services'!{src}{bref}")
        cc.font, cc.number_format = DERIVED, MONEY
    ps.cell(row=pr, column=5, value=STATUS_WORD.get(b["build_status"])).font = BODY
    pr += 1

    first_part = pr
    for p in parts:
        pref = row_of[p["code"]]
        ps.cell(row=pr, column=1, value="    " + p["name"]).font = BODY
        ps.cell(row=pr, column=2, value=p["code"]).font = MUTED
        for col, src in ((3, "E"), (4, "F")):
            cc = ps.cell(row=pr, column=col, value=f"='All Services'!{src}{pref}")
            cc.font, cc.number_format = BODY, MONEY
        pr += 1
    last_part = pr - 1

    ps.cell(row=pr, column=1, value="    Bought separately").font = BODY
    for col in (3, 4):
        L = get_column_letter(col)
        cc = ps.cell(row=pr, column=col, value=f"=SUM({L}{first_part}:{L}{last_part})")
        cc.font, cc.number_format = BODY, MONEY
    sep_row = pr
    pr += 1

    ps.cell(row=pr, column=1, value="    You save").font = DERIVED
    for col in (3, 4):
        L = get_column_letter(col)
        cc = ps.cell(row=pr, column=col, value=f"={L}{sep_row}-{L}{bundle_row}")
        cc.font, cc.number_format = DERIVED, MONEY
    # percentage saved, guarded against a zero denominator
    for col, tgt in ((5, 3), (6, 4)):
        L = get_column_letter(tgt)
        cc = ps.cell(row=pr, column=col,
                     value=f"=IF({L}{sep_row}=0,0,({L}{sep_row}-{L}{bundle_row})/{L}{sep_row})")
        cc.font, cc.number_format = DERIVED, PCT
    ps.cell(row=pr, column=5).comment = Comment(
        "Setup % saved. Packages are priced off the individual prices, never off "
        "another package, so a discount is never applied twice.", "Sanaku")
    pr += 2

# ---------------------------------------------------------------------------
# 3. One tab per vertical.
# ---------------------------------------------------------------------------
for key, label in (("home_services", "Home Services"), ("law_firm", "Law Firms"), ("medical", "Medical")):
    vs = wb.create_sheet(label)
    vs["A1"] = f"{label} — what this vertical can buy"
    vs["A1"].font = H1
    header(vs, 3, ["Service", "Code", "Type", "Setup", "Monthly", "Per lead", "Status"],
           [32, 26, 10, 12, 12, 12, 16])
    vr = 4
    for s in services:
        if key not in s["allowed_verticals"]:
            continue
        ref = row_of[s["code"]]
        vs.cell(row=vr, column=1, value=s["name"]).font = BODY
        vs.cell(row=vr, column=2, value=s["code"]).font = MUTED
        vs.cell(row=vr, column=3, value=s["category"]).font = BODY
        for col, src in ((4, "E"), (5, "F"), (6, "J")):
            cc = vs.cell(row=vr, column=col, value=f"='All Services'!{src}{ref}")
            cc.font, cc.number_format = BODY, MONEY
        vs.cell(row=vr, column=7, value=STATUS_WORD.get(s["build_status"])).font = BODY
        vr += 1
    if key == "medical":
        vs.cell(row=vr + 1, column=1,
                value="Every medical service requires a signed BAA covering every vendor in "
                      "the path before go-live, and is flat-fee only.").font = MUTED
        vs.cell(row=vr + 1, column=1).fill = NOTE_FILL

# ---------------------------------------------------------------------------
# 4. Scenarios — first-year revenue per client. Revenue, not margin.
# ---------------------------------------------------------------------------
sc = wb.create_sheet("Scenarios")
sc["A1"] = "What one client is worth in year one"
sc["A1"].font = H1
sc["A2"] = ("Setup plus twelve months. Edit the client count in blue to size a book of business. "
            "Revenue only — no cost data exists in the system, so there is no margin here to show.")
sc["A2"].font = MUTED
sc.merge_cells("A2:F2")

header(sc, 4, ["Package", "Clients", "Setup each", "Monthly each", "Year one each", "Year one total"],
       [34, 10, 14, 14, 16, 18])

sr = 5
for b in bundles:
    pref = None
    for i in range(5, pr):
        if ps.cell(row=i, column=2).value == b["code"]:
            pref = i
            break
    ref = row_of[b["code"]]
    sc.cell(row=sr, column=1, value=f'{b["name"]} — {", ".join(b["allowed_verticals"])}').font = BODY
    c = sc.cell(row=sr, column=2, value=1)
    c.font = INPUT
    for col, src in ((3, "E"), (4, "F")):
        cc = sc.cell(row=sr, column=col, value=f"='All Services'!{src}{ref}")
        cc.font, cc.number_format = BODY, MONEY
    cc = sc.cell(row=sr, column=5, value=f"=C{sr}+(D{sr}*12)")
    cc.font, cc.number_format = DERIVED, MONEY
    cc = sc.cell(row=sr, column=6, value=f"=B{sr}*E{sr}")
    cc.font, cc.number_format = DERIVED, MONEY
    sr += 1

sc.cell(row=sr, column=1, value="Total").font = DERIVED
cc = sc.cell(row=sr, column=6, value=f"=SUM(F5:F{sr - 1})")
cc.font, cc.number_format = DERIVED, MONEY
sc.cell(row=sr + 2, column=1,
        value="Excludes per-lead fees and any usage beyond the included allowance, both of which "
              "vary by client. Year one only — setup is not repeated in year two.").font = MUTED
sc.cell(row=sr + 2, column=1).fill = NOTE_FILL

out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(out_path)
print(f"wrote {out_path}")
print(f"  {len(services)} services, {len(bundles)} packages, {len(wb.sheetnames)} tabs")
